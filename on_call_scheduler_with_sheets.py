import pandas as pd
from collections import defaultdict
import random
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import calendar
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# Developer color mapping (RGB values 0-1 for Google Sheets API)
DEVELOPER_COLORS = {
    "Omer": {"red": 0.6, "green": 0.7, "blue": 0.9},           # Blue
    "Shlomi": {"red": 0.7, "green": 0.9, "blue": 0.9},         # Light cyan
    "Amit": {"red": 0.85, "green": 0.75, "blue": 0.6},         # Light brown/tan
    "Ohad": {"red": 1.0, "green": 0.8, "blue": 0.8},           # Light pink
    "Ivan": {"red": 0.95, "green": 0.6, "blue": 0.6},          # Red/coral
    "Gabriel": {"red": 0.7, "green": 0.9, "blue": 0.7},        # Light green
    "Or": {"red": 0.85, "green": 0.7, "blue": 0.9},            # Purple/lavender
    "Hagay": {"red": 0.7, "green": 0.5, "blue": 0.8},          # Darker purple
    "Yariv": {"red": 0.8, "green": 0.4, "blue": 0.4},          # Dark red/maroon
    "Alex": {"red": 0.9, "green": 0.9, "blue": 0.6},           # Light yellow
}

# Load month/year and developer constraints from JSON file
def load_data_from_json(json_file="data/constraints.json"):
    """Load month, year, and developer constraints from JSON file"""
    try:
        with open(json_file, 'r') as f:
            data = json.load(f)

            # Get month and year (defaults to next month if not specified)
            now = datetime.now()
            next_month = datetime(now.year, now.month, 1)
            if now.month == 12:
                next_month = datetime(now.year + 1, 1, 1)
            else:
                next_month = datetime(now.year, now.month + 1, 1)

            month = data.get("month", next_month.month)
            year = data.get("year", next_month.year)

            # Get developers and their restrictions
            developers = {}
            for dev_name, dev_data in data.get("developers", {}).items():
                developers[dev_name] = dev_data.get("restrictions", [])

            return month, year, developers
    except FileNotFoundError:
        print(f"Warning: {json_file} not found. Using hardcoded constraints.")
        return None, None, None

# Try to load from JSON
month, year, developers = load_data_from_json()

# Get month name for display
month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

if month is None or year is None or developers is None:
    # Fallback to hardcoded values
    print("⚠️  Using hardcoded month and constraints")
    month = 2
    year = 2026

else:
    print(f"✓ Planning schedule for: {month_names[month-1]} {year}")

if developers is None or not developers:
    # Fallback: Developers and their restrictions (hardcoded)
    developers = {
        "Gabriel": [
            "12/02 Day", "13/02 Night", "13/02 Day", "22/02 Day", "23/02 Night", "23/02 Day", "24/02 Night"
        ],
        "Shlomi": [],
        "Yariv": [],
        "Omer": [],
        "Amit": [
            "06/02 Day", "25/02 Day"
        ],
        "Ohad": [
            "01/02 Day", "01/02 Night", "02/02 Day", "02/02 Night", "03/02 Day", "03/02 Night",
            "04/02 Day", "04/02 Night", "17/02 Day", "17/02 Night", "18/02 Day", "18/02 Night"
        ],
        "Alex": [
            "01/02 Night", "01/02 Day", "02/02 Day", "02/02 Night", "03/02 Day", "10/02 Day",
            "17/02 Day", "24/02 Day", "27/02 Day", "27/02 Night", "28/02 Day", "28/02 Night"
        ],
        "Or": [
            "02/02 Day", "05/02 Day", "06/02 Day", "06/02 Night", "07/02 Day", "07/02 Night",
            "09/02 Day", "13/02 Day", "23/02 Day"
        ],
        "Hagay": [],
        "Ivan": []
    }

# Generate all days of the month based on loaded month/year
last_day = calendar.monthrange(year, month)[1]
start_date = f"{year}-{month:02d}-01"
end_date = f"{year}-{month:02d}-{last_day}"

print(f"📅 Generating schedule from {start_date} to {end_date}")

all_days = pd.date_range(start=start_date, end=end_date, freq="D")
shifts_data = [
    {"date": day.strftime("%d/%m"), "day_of_week": day.strftime("%A")}
    for day in all_days
]

# Create two shifts (Day and Night) for each day
expanded_shifts_data = []
for day in shifts_data:
    expanded_shifts_data.append({**day, "shift": "Day"})
    expanded_shifts_data.append({**day, "shift": "Night"})

# Special shifts
special_shifts = {"Friday Day", "Friday Night", "Saturday Day", "Saturday Night"}
# Israeli holidays - add dates here in format "DD/MM" (e.g., "14/04" for April 14)
events = {
    # 2026 Israeli Holidays
    "14/04 Day", "14/04 Night",  # Passover
    "15/04 Day", "15/04 Night",  # Passover
    "04/05 Day", "04/05 Night",  # Independence Day
    "05/05 Day", "05/05 Night",  # Independence Day
    "25/05 Day", "25/05 Night",  # Shavuot
    "13/09 Day", "13/09 Night",  # Rosh Hashanah
    "14/09 Day", "14/09 Night",  # Rosh Hashanah
    "22/09 Day", "22/09 Night",  # Yom Kippur
    "27/09 Day", "27/09 Night",  # Sukkot
    "28/09 Day", "28/09 Night",
    "29/09 Day", "29/09 Night",
    "30/09 Day", "30/09 Night",
    "01/10 Day", "01/10 Night",
    "02/10 Day", "02/10 Night",
    "03/10 Day", "03/10 Night",
    "04/10 Day", "04/10 Night",  # Sukkot
}

# Convert to DataFrame
shifts_df = pd.DataFrame(expanded_shifts_data)

# Initialize assignment tracking
assignments = defaultdict(list)
developer_shift_count = defaultdict(lambda: {"special": 0, "night": 0, "day": 0})


# Calculate totals
total_night_shifts = sum(1 for shift in expanded_shifts_data if shift["shift"] == "Night")
total_special_shifts = sum(1 for shift in expanded_shifts_data if f'{shift["day_of_week"]} {shift["shift"]}' in special_shifts or f'{shift["date"]} {shift["shift"]}' in events)

developers_list = list(developers.keys())
# Shuffle the list to prevent order bias
random.shuffle(developers_list)
night_shifts_per_dev = total_night_shifts // len(developers_list) +1
special_shifts_per_dev = total_special_shifts // len(developers_list) +1

remaining_night = total_night_shifts % len(developers_list)
remaining_special = total_special_shifts % len(developers_list)


# Function to get the least assigned developer (prioritizing shift type balance)
def get_least_assigned(shift_type, eligible_developers):
    # Sort by specific shift type first, then by total shifts as tiebreaker
    sorted_devs = sorted(
        eligible_developers,
        key=lambda dev: (developer_shift_count[dev][shift_type], sum(developer_shift_count[dev].values()))
    )

    # If multiple developers have the same shift type count, randomly select among them
    min_shift_type = developer_shift_count[sorted_devs[0]][shift_type]
    tied_devs = [dev for dev in sorted_devs if developer_shift_count[dev][shift_type] == min_shift_type]

    if len(tied_devs) > 1:
        return random.choice(tied_devs)
    else:
        return sorted_devs[0]

# Function to check if a developer can take a shift
def can_assign(developer, date, shift, last_assigned):
    # Check restrictions
    if f"{date} {shift}" in developers[developer]:
        return False
    # Avoid consecutive assignments
    if last_assigned and last_assigned == developer:
        return False
    if (f"{date} {shift}" in events or f"{day_of_week} {shift}" in special_shifts) and developer_shift_count[developer]["special"] >= special_shifts_per_dev:
        return False
    if shift == "Night" and developer_shift_count[developer]["night"] >= night_shifts_per_dev:
        return False
    return True


# Assign shifts
last_assigned = None
assigned_shifts = []
for _, row in shifts_df.iterrows():
    date, shift, day_of_week = row["date"], row["shift"], row["day_of_week"]
    shift_type = "special" if (f"{day_of_week} {shift}" in special_shifts) or ( f"{date} {shift}" in events) else ("night" if shift == "Night" else "day")
    eligible_developers = [
        dev for dev in developers_list
        if can_assign(dev, date, shift, last_assigned)
    ]

    if eligible_developers:
        assigned_dev = get_least_assigned(shift_type, eligible_developers)
    else:
        assigned_dev = None

    if assigned_dev:
        developer_shift_count[assigned_dev][shift_type] += 1
        last_assigned = assigned_dev

    assigned_shifts.append({
        "day_of_month": int(date.split("/")[0]),
        "night_shift": assigned_dev if shift == "Night" else None,
        "day_shift": assigned_dev if shift == "Day" else None,
        "day_of_week": day_of_week
    })

# Combine assigned shifts into a single row per day
final_assignments = []
temp_row = {}
for shift in assigned_shifts:
    day_of_month = shift["day_of_month"]
    if temp_row.get("day_of_month") != day_of_month:
        if temp_row:
            final_assignments.append(temp_row)
        temp_row = {
            "day_of_month": day_of_month,
            "night_shift": None,
            "day_shift": None,
            "day_of_week": shift["day_of_week"]
        }
    if shift["night_shift"]:
        temp_row["night_shift"] = shift["night_shift"]
    if shift["day_shift"]:
        temp_row["day_shift"] = shift["day_shift"]

# Append the last row
if temp_row:
    final_assignments.append(temp_row)

# Write to CSV
final_df = pd.DataFrame(final_assignments)
# Rename columns to match template format
final_df = final_df.rename(columns={
    'day_of_month': 'Day of Month',
    'night_shift': 'First Assignment',
    'day_shift': 'Second Assignment',
    'day_of_week': 'Day of Week'
})

# Create output directory if it doesn't exist
output_dir = "output"
os.makedirs(output_dir, exist_ok=True)

# Add month header row to match template format
month_header = pd.DataFrame([[month_names[month-1], '', '', '']], columns=final_df.columns)
final_df_with_header = pd.concat([month_header, final_df], ignore_index=True)

# Define output file paths
schedule_file = os.path.join(output_dir, "shift_schedule.csv")
summary_file = os.path.join(output_dir, "shift_summary.csv")

# Write schedule to CSV
final_df_with_header.to_csv(schedule_file, index=False)

# Print summary of shifts
print("\nShift Summary:")
summary_df = pd.DataFrame(developer_shift_count).T.reset_index()
summary_df.columns = ["Developer", "Special Shifts", "Night Shifts", "Day Shifts"]
print(summary_df)

# Write summary to CSV
summary_df.to_csv(summary_file, index=False)

print(f"\n✓ Shift schedule written to {schedule_file}")
print(f"✓ Shift summary written to {summary_file}")

# ============================================================================
# EXCEL (XLSX) EXPORT WITH COLORS
# ============================================================================

def create_excel_with_colors(df, month_name, output_path):
    """
    Create an Excel file with the same color formatting as Google Sheets

    Args:
        df: DataFrame with schedule data (without header row)
        month_name: Month name (e.g., "Mar")
        output_path: Path to save the Excel file
    """
    print("\n" + "="*60)
    print("Creating Excel file with colors...")
    print("="*60)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Write month name in A1 with bold formatting
    ws['A1'] = month_name
    ws['A1'].font = Font(bold=True, size=12)

    # Write headers in row 2
    headers = ['Day of Month', 'Night Shift', 'Day Shift', 'Day of Week']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data starting from row 3
    for idx, row_data in df.iterrows():
        row_num = idx + 3  # Start from row 3 (1=month, 2=headers)

        # Column A: Day of Month
        ws.cell(row=row_num, column=1, value=int(row_data['Day of Month']))

        # Column B: Night Shift (with color)
        night_dev = row_data['First Assignment']
        night_cell = ws.cell(row=row_num, column=2, value=night_dev)
        night_cell.alignment = Alignment(horizontal='center')

        if night_dev and night_dev in DEVELOPER_COLORS:
            color = DEVELOPER_COLORS[night_dev]
            # Convert RGB 0-1 to hex color (0-255)
            r = int(color['red'] * 255)
            g = int(color['green'] * 255)
            b = int(color['blue'] * 255)
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            night_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # Column C: Day Shift (with color)
        day_dev = row_data['Second Assignment']
        day_cell = ws.cell(row=row_num, column=3, value=day_dev)
        day_cell.alignment = Alignment(horizontal='center')

        if day_dev and day_dev in DEVELOPER_COLORS:
            color = DEVELOPER_COLORS[day_dev]
            # Convert RGB 0-1 to hex color (0-255)
            r = int(color['red'] * 255)
            g = int(color['green'] * 255)
            b = int(color['blue'] * 255)
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            day_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # Column D: Day of Week
        ws.cell(row=row_num, column=4, value=row_data['Day of Week'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Save the workbook
    wb.save(output_path)
    print(f"✓ Excel file with colors written to {output_path}")
    print("="*60)

# Create Excel file with colors
excel_file = os.path.join(output_dir, "shift_schedule.xlsx")
create_excel_with_colors(final_df, month_names[month-1], excel_file)


# ============================================================================
# GOOGLE SHEETS INTEGRATION
# ============================================================================

def upload_to_google_sheets(df, config, year, month_name):
    """
    Append schedule to Google Sheets with developer color coding

    Args:
        df: DataFrame with schedule data (without header row)
        config: Configuration dict with spreadsheet_id, worksheet_name, credentials_file
        year: Year for the schedule
        month_name: Month name (e.g., "Feb")
    """
    try:
        print("\n" + "="*60)
        print("Uploading to Google Sheets...")
        print("="*60)

        # Setup credentials
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            config['credentials_file'], scope
        )
        client = gspread.authorize(creds)

        # Open spreadsheet by name "On Call Schedule"
        spreadsheet = client.open("On Call Schedule")
        print(f"✓ Connected to spreadsheet: {spreadsheet.title}")

        # Get or create worksheet with name "On call schedule YYYY"
        worksheet_name = f"On call schedule {year}"
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
            print(f"✓ Found worksheet: {worksheet_name}")
        except:
            worksheet = spreadsheet.add_worksheet(
                title=worksheet_name,
                rows=500,
                cols=10
            )
            print(f"✓ Created new worksheet: {worksheet_name}")

        # Find the next empty row (append mode)
        all_values = worksheet.get_all_values()
        next_row = len(all_values) + 1

        # If sheet is empty, start from row 1
        if next_row == 1:
            next_row = 1
        else:
            # Add 2 empty rows
            next_row += 2

        print(f"✓ Appending data starting at row {next_row}")

        # Write month name in column A
        month_row = next_row
        worksheet.update_cell(month_row, 1, month_name)  # Column A
        print(f"✓ Wrote month name '{month_name}' at row {month_row}")

        # Prepare schedule data (A=day, B=night, C=day, D=day_of_week)
        schedule_start_row = month_row + 1
        data_rows = []

        for _, row in df.iterrows():
            data_rows.append([
                row['Day of Month'],      # Column A: Day number
                row['First Assignment'],  # Column B: Night shift
                row['Second Assignment'], # Column C: Day shift
                row['Day of Week']        # Column D: Day of week
            ])

        # Upload schedule data
        if data_rows:
            cell_range = f'A{schedule_start_row}:D{schedule_start_row + len(data_rows) - 1}'
            worksheet.update(cell_range, data_rows)
            print(f"✓ Uploaded {len(data_rows)} days of schedule")

        # Apply formatting
        print("\nApplying formatting...")

        # Format requests for batch update
        requests = []

        # Bold and larger font for month name
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet.id,
                    "startRowIndex": month_row - 1,
                    "endRowIndex": month_row,
                    "startColumnIndex": 0,
                    "endColumnIndex": 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True, "fontSize": 12}
                    }
                },
                "fields": "userEnteredFormat.textFormat"
            }
        })

        # Apply colors and center alignment to developer cells (columns B and C)
        for idx, row in enumerate(df.iterrows()):
            _, row_data = row
            row_idx = schedule_start_row - 1 + idx  # Zero-indexed for API

            # Night shift (column B, index 1)
            night_dev = row_data['First Assignment']
            if night_dev and night_dev in DEVELOPER_COLORS:
                color = DEVELOPER_COLORS[night_dev]
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": worksheet.id,
                            "startRowIndex": row_idx,
                            "endRowIndex": row_idx + 1,
                            "startColumnIndex": 1,  # Column B
                            "endColumnIndex": 2
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": color,
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment)"
                    }
                })

            # Day shift (column C, index 2)
            day_dev = row_data['Second Assignment']
            if day_dev and day_dev in DEVELOPER_COLORS:
                color = DEVELOPER_COLORS[day_dev]
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": worksheet.id,
                            "startRowIndex": row_idx,
                            "endRowIndex": row_idx + 1,
                            "startColumnIndex": 2,  # Column C
                            "endColumnIndex": 3
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": color,
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment)"
                    }
                })

        # Execute batch update
        if requests:
            spreadsheet.batch_update({"requests": requests})
            print(f"✓ Applied colors to {len(requests)} cells")

        print("\n" + "="*60)
        print(f"✓ SUCCESS! Schedule uploaded to Google Sheets")
        print(f"  Spreadsheet: {spreadsheet.title}")
        print(f"  Worksheet: {worksheet_name}")
        print(f"  URL: {spreadsheet.url}")
        print("="*60 + "\n")

    except Exception as e:
        print(f"\n⚠ Warning: Could not upload to Google Sheets")
        print(f"  Error: {e}")
        print(f"  CSV files saved locally in output/ folder\n")


# Upload to Google Sheets if configuration exists and enabled
try:
    with open('config.json', 'r') as f:
        config = json.load(f)

    # Check if upload is enabled (default to True if not specified)
    upload_enabled = config.get('upload_to_sheets', True)

    if upload_enabled:
        # Pass dataframe without header, year, and month name
        upload_to_google_sheets(final_df, config, year, month_names[month-1])
    else:
        print("\n" + "="*60)
        print("ℹ️  Google Sheets upload is DISABLED in config.json")
        print("   Schedule saved locally in output/ folder")
        print("   To enable: Set 'upload_to_sheets': true in config.json")
        print("="*60 + "\n")

except FileNotFoundError:
    print("\n⚠ config.json not found. Skipping Google Sheets upload.")
    print("  To enable Google Sheets integration:")
    print("  1. Create config.json with your spreadsheet details")
    print("  2. Set up Google Sheets API credentials")
    print("  See docs/SETUP.md for instructions.\n")
except Exception as e:
    print(f"\n⚠ Error loading config: {e}\n")
